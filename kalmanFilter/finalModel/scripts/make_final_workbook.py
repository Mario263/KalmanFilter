"""Build FINAL_RAW_VS_KALMAN_STUDY.xlsx from the finalModel JSON/CSV artifacts.

Reads only already-generated artifacts (no recompute, no fabricated values). Sheets:
README, Final_Config, Raw_Metrics, Raw_Nautilus, Kalman_Metrics, Kalman_Nautilus,
Ensemble_Metrics, Final_Comparison, Seed_Dependence, Verification_Gates, Artifacts,
Commands. Top row frozen, auto-filter + auto-size on every sheet.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
FINAL = HERE.parent
OUT = FINAL / "outputs"
WB = FINAL / "workbook" / "FINAL_RAW_VS_KALMAN_STUDY.xlsx"

METRIC_KEYS = ["cumulative_return", "cagr", "sharpe", "sortino", "calmar", "max_drawdown",
               "var_95", "win_rate", "active_win_rate", "trade_win_rate", "round_trips",
               "final_equity", "total_turnover", "long_frac", "flat_frac", "short_frac"]


def _j(p):
    p = Path(p)
    return json.loads(p.read_text()) if p.exists() else None


def _sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    last = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last}{max(ws.max_row, 1)}"
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        width = max([len(str(headers[c - 1]))] + [len(str(r[c - 1])) for r in rows if c - 1 < len(r)] + [6])
        ws.column_dimensions[col].width = min(width + 2, 60)
    return ws


def _metric_rows(label, block):
    if not block:
        return [[label] + ["n/a"] * len(METRIC_KEYS)]
    return [[label] + [round(block.get(k), 6) if isinstance(block.get(k), float) else block.get(k) for k in METRIC_KEYS]]


def build():
    spec = _j(FINAL / "configs" / "final_model_spec.json") or {}
    raw_s = _j(OUT / "raw" / "metrics" / "eval_s23.json")
    kal_s = _j(OUT / "kalman" / "metrics" / "eval_s23.json")
    raw_e = _j(OUT / "raw" / "metrics" / "eval_ensemble.json")
    kal_e = _j(OUT / "kalman" / "metrics" / "eval_ensemble.json")
    raw_n = _j(OUT / "raw" / "metrics" / "nautilus_ppo_raw_s23.json")
    kal_n = _j(OUT / "kalman" / "metrics" / "nautilus_ppo_kalman_s23.json")
    cmp_ = _j(OUT / "comparison" / "final_comparison.json") or {}
    gate = _j(OUT / "verification" / "final_gate_result.json") or {}
    summ = _j(OUT / "verification" / "seed_dependence_summary.json") or {}

    wb = Workbook()
    wb.remove(wb.active)

    # README
    _sheet(wb, "README", ["Section", "Description"], [
        ["Purpose", "Final productionized Raw PPO vs Kalman PPO; robust seed-ensemble + paper-faithful seed-23."],
        ["Final selected configuration", "exp_002_final: OHLC-4D full-covariance EM Kalman, raw-close PnL, 22-D state."],
        ["Raw comparator", "rawctl: true Raw PPO (no Kalman), identical frozen PPO/reward/costs/split."],
        ["Kalman configuration", "filtered OHLC + raw Volume observation; full covariance; EM train-only Q/R; no transform."],
        ["Price-basis rule", "PnL/reward/equity/Nautilus fills ALL use raw close. price/raw_close never in the 22-D obs."],
        ["Seed-dependence fix", f"soft-vote ensemble of seeds {spec.get('ensemble_seeds')} (deployed model is seed-independent)."],
        ["Outperformance gate", "Kalman Sharpe > Raw AND Kalman MaxDD less severe (evaluated on the ensemble)."],
        ["Caveat", "Edge is modest and risk-adjusted; single seed-23 may not beat a lucky historical raw checkpoint."],
    ])

    # Final_Config (flattened)
    cfg_rows = []
    def flat(prefix, d):
        for k, v in d.items():
            key = f"{prefix}{k}"
            if isinstance(v, dict):
                flat(key + ".", v)
            else:
                cfg_rows.append([key, str(v)])
    flat("", spec)
    _sheet(wb, "Final_Config", ["Field", "Value"], cfg_rows)

    # per-family metric sheets
    _sheet(wb, "Raw_Metrics", ["source"] + METRIC_KEYS, _metric_rows("raw_seed23_env", (raw_s or {}).get("kalman_env")))
    _sheet(wb, "Raw_Nautilus", ["source"] + METRIC_KEYS + ["min_equity", "nonpositive_periods"],
           [_metric_rows("raw_seed23_nautilus", (raw_n or {}).get("nautilus"))[0] +
            [(raw_n or {}).get("nautilus", {}).get("min_equity"), (raw_n or {}).get("nautilus", {}).get("nonpositive_periods")]])
    _sheet(wb, "Kalman_Metrics", ["source"] + METRIC_KEYS, _metric_rows("kalman_seed23_env", (kal_s or {}).get("kalman_env")))
    _sheet(wb, "Kalman_Nautilus", ["source"] + METRIC_KEYS + ["min_equity", "nonpositive_periods"],
           [_metric_rows("kalman_seed23_nautilus", (kal_n or {}).get("nautilus"))[0] +
            [(kal_n or {}).get("nautilus", {}).get("min_equity"), (kal_n or {}).get("nautilus", {}).get("nonpositive_periods")]])

    # Ensemble metrics (the robust production result)
    ens_rows = (_metric_rows("raw_ensemble", (raw_e or {}).get("kalman_env")) +
                _metric_rows("kalman_ensemble", (kal_e or {}).get("kalman_env")))
    _sheet(wb, "Ensemble_Metrics", ["source"] + METRIC_KEYS, ens_rows)

    # Final_Comparison (ensemble + single seed-23)
    cmp_rows = []
    for basis in ("ensemble", "single_seed23"):
        for r in cmp_.get(basis, []):
            cmp_rows.append([basis, r["metric"], r["raw_value"], r["kalman_value"], r["delta"], r["winner"]])
    _sheet(wb, "Final_Comparison", ["basis", "metric", "raw_value", "kalman_value", "delta", "winner"], cmp_rows)

    # Seed_Dependence
    sd_rows = []
    sdc = OUT / "verification" / "seed_dependence_diagnostics.csv"
    if sdc.exists():
        with open(sdc, newline="") as f:
            rdr = list(csv.reader(f))
        hdr = rdr[0]
        sd_rows = rdr[1:]
    else:
        hdr = ["family", "seed", "cumulative_return", "sharpe", "max_drawdown", "round_trips", "exposure", "final_equity"]
    _sheet(wb, "Seed_Dependence", hdr, sd_rows)

    # Verification_Gates
    g_rows = []
    for k, v in (gate.get("gates") or {}).items():
        g_rows.append([k, "PASS" if v else "FAIL", "", "outputs/verification/final_gate_result.json"])
    if not g_rows:
        g_rows = [["(run verify stage)", "n/a", "", "outputs/verification/final_gate_result.json"]]
    _sheet(wb, "Verification_Gates", ["gate", "status", "details", "artifact_path"], g_rows)

    # Artifacts
    art = [
        ["final_model_spec", "configs/final_model_spec.json"],
        ["raw_eval_seed23", "outputs/raw/metrics/eval_s23.json"],
        ["kalman_eval_seed23", "outputs/kalman/metrics/eval_s23.json"],
        ["raw_ensemble_eval", "outputs/raw/metrics/eval_ensemble.json"],
        ["kalman_ensemble_eval", "outputs/kalman/metrics/eval_ensemble.json"],
        ["raw_nautilus_seed23", "outputs/raw/metrics/nautilus_ppo_raw_s23.json"],
        ["kalman_nautilus_seed23", "outputs/kalman/metrics/nautilus_ppo_kalman_s23.json"],
        ["final_comparison", "outputs/comparison/final_comparison.json"],
        ["final_gate_result", "outputs/verification/final_gate_result.json"],
        ["indicator_basis", "outputs/verification/indicator_basis_verification.json"],
        ["seed_dependence_summary", "outputs/verification/seed_dependence_summary.json"],
        ["seed_dependence_csv", "outputs/verification/seed_dependence_diagnostics.csv"],
    ]
    art_rows = [[t, p, (FINAL / p).exists(), ""] for t, p in art]
    _sheet(wb, "Artifacts", ["artifact_type", "path", "exists", "description"], art_rows)

    # Commands
    base = r'.\.venv\Scripts\python.exe kalmanFilter\finalModel\scripts\run_final_model.py'
    _sheet(wb, "Commands", ["task", "command", "notes"], [
        ["run everything", f"{base} --stage all --device cpu", "train+eval+nautilus+diagnose+compare+verify+excel"],
        ["train raw", f"{base} --stage train_raw --device cpu", "5 seeds"],
        ["train kalman", f"{base} --stage train_kalman --device cpu", "5 seeds"],
        ["compare", f"{base} --stage compare --device cpu", "writes final_comparison.*"],
        ["verify", f"{base} --stage verify --device cpu", "writes final_gate_result.json"],
        ["excel", f"{base} --stage excel --device cpu", "rebuild this workbook"],
    ])

    WB.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WB)
    # validate
    from openpyxl import load_workbook
    load_workbook(WB)
    print(f"[excel] wrote {WB} | sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    build()
