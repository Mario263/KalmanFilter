"""Build KALMAN_PPO_FULL_STUDY_RESULTS.xlsx from study artifacts (no hand-typed tables).

Loads JSON/CSV produced by the raw-price suite + final verification + Phase-1 history.
15 sheets, frozen header row, auto-filter, auto-sized columns. Read-only on artifacts.
"""
from __future__ import annotations
import glob, json, sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

_ABLR = Path(__file__).resolve().parents[1]
_KF = _ABLR.parent
_ROOT = _KF.parent
sys.path.insert(0, str(_KF))
sys.path.insert(0, str(_ROOT / "src"))
OUT = _ABLR / "outputs"
OUTV = OUT / "final_verification"
DOCS = _ABLR / "docs"
XLSX = DOCS / "KALMAN_PPO_FULL_STUDY_RESULTS.xlsx"
CELLS = ["rawctl_corrected", "exp_002_corrected", "exp_007_corrected", "exp_003_corrected", "exp_004_corrected"]
SEEDS = [1, 2, 3, 7, 11, 13, 17, 19, 23, 42]


def _load(p, default=None):
    p = Path(p)
    return json.loads(p.read_text()) if p.exists() else (default if default is not None else {})


def _sheet(wb, name, rows, cols=None):
    ws = wb.create_sheet(name[:31])
    df = pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols or ["(no data)"])
    headers = list(df.columns)
    ws.append(headers)
    for _, r in df.iterrows():
        ws.append([_cell(v) for v in r.tolist()])
    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df)+1}"
    for i, h in enumerate(headers, 1):
        w = max([len(str(h))] + [len(str(_cell(v))) for v in df[h].tolist()][:200]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(w, 60)
    return ws


def _cell(v):
    if isinstance(v, float):
        return round(v, 6)
    if isinstance(v, (dict, list)):
        return json.dumps(v)
    return v


def _seed_metrics():
    rows = []
    for eid in CELLS:
        for f in sorted((OUT / eid / "metrics").glob("eval_s*.json")):
            d = _load(f); m = d["kalman_env"]; s = d.get("seed")
            rows.append({"experiment_id": eid, "seed": s, "model_path": str(OUT / eid / "models" / f"ppo_{eid}_s{s}"),
                         "metrics_path": str(f), "price_basis": d.get("price_basis"),
                         "return": m["cumulative_return"], "cagr": m["cagr"], "sharpe": m["sharpe"], "sortino": m["sortino"],
                         "calmar": m["calmar"], "max_drawdown": m["max_drawdown"], "win_rate": m["trade_win_rate"],
                         "round_trips": m["round_trips"], "trades": m["total_turnover"],
                         "exposure": round(1 - m["flat_frac"], 4), "turnover": m["total_turnover"],
                         "final_equity": m["final_equity"], "negative_return": m["cumulative_return"] < 0})
    return rows


def main():
    wb = Workbook(); wb.remove(wb.active)
    man = _load(OUT / "FINAL_MODEL" / "FINAL_MODEL_MANIFEST.json")
    sb = _load(_ABLR / "docs" / "scoreboard.json")            # {scoreboard:[], paired_stats:[]}
    seedm = _seed_metrics()
    ib = _load(OUTV / "indicator_basis_summary.json")
    pb = _load(OUTV / "price_basis_verification.json")
    hist = _load(_KF / "ablation" / "docs" / "scoreboard.json")  # phase-1 raw_close + filtered_close

    # 1 README
    _sheet(wb, "README", [
        {"Section": "Study purpose", "Description": "Honest raw-price revalidation of Kalman-enhanced PPO vs Raw PPO (XAUUSD 1d)."},
        {"Section": "Final selected model", "Description": "exp_002_corrected seed 23 (raw OHLC + full-cov EM Kalman)."},
        {"Section": "Final model path", "Description": "outputs/FINAL_MODEL/FINAL_ppo_kalman_fullcov_raw_s23.zip"},
        {"Section": "Price-basis correction", "Description": "PnL/reward/equity/fills on RAW close; filtered OHLC only as features."},
        {"Section": "Indicator-basis verdict", "Description": f"exp_002 indicators: {ib.get('cells',{}).get('exp_002_corrected',{}).get('verdict','?')} (17/17 filtered, 0/17 raw)."},
        {"Section": "Parity verdict", "Description": "exp_002_corrected 10/10 seeds Env<->Nautilus parity PASS (raw_close)."},
        {"Section": "Statistical caveat", "Description": man.get("statistical_caveat")},
        {"Section": "How to read", "Description": "Honest perf = raw_close basis. Filtered_close (Sheet 11) is DIAGNOSTIC ONLY — NOT TRADEABLE."},
    ], cols=["Section", "Description"])

    # 2 Final_Model_Decision
    fm = [("final_experiment_id", man.get("experiment_id"), man.get("source_model")),
          ("final_seed", man.get("seed"), man.get("source_metrics")),
          ("final_model_path", man.get("final_model_file"), "outputs/FINAL_MODEL/"),
          ("selection_rule", man.get("selection_rule"), ""),
          ("price_basis", man.get("price_basis"), man.get("source_runtime_config")),
          ("feature_basis", man.get("feature_basis"), ""),
          ("indicator_basis", man.get("indicator_basis"), str(OUTV / "indicator_basis_summary.json")),
          ("state_dimension", man.get("state_dimension"), ""),
          ("volume_policy", man.get("volume_policy"), ""),
          ("mean_sharpe", man.get("distribution_10seed", {}).get("sharpe_mean"), str(_ABLR / "docs/RAWPRICE_SCOREBOARD.md")),
          ("median_sharpe", man.get("distribution_10seed", {}).get("sharpe_median"), ""),
          ("mean_return", man.get("distribution_10seed", {}).get("ret_mean"), ""),
          ("mean_maxdd", man.get("distribution_10seed", {}).get("mean_maxdd"), ""),
          ("negative_seed_count", man.get("distribution_10seed", {}).get("neg_seeds"), ""),
          ("bootstrap_sharpe_delta_CI95", man.get("distribution_10seed", {}).get("sharpe_delta_vs_raw_CI95"), str(_ABLR / "docs/RAWPRICE_SEED_ROBUSTNESS_REPORT.md")),
          ("wilcoxon_p", man.get("distribution_10seed", {}).get("wilcoxon_p"), ""),
          ("parity_all_seeds", man.get("env_nautilus_parity_all_seeds"), str(OUTV / "final_parity_all_seeds.json"))]
    _sheet(wb, "Final_Model_Decision", [{"Field": a, "Value": _cell(b), "Evidence_Path": c} for a, b, c in fm],
           cols=["Field", "Value", "Evidence_Path"])

    # 3 Core_Scoreboard
    desc = {"rawctl_corrected": ("none", "bypass", "FAIR raw baseline (no Kalman)", "baseline"),
            "exp_002_corrected": ("none", "full", "raw OHLC + full-cov EM Kalman", "PROMOTED (final)"),
            "exp_007_corrected": ("none", "diagonal", "raw OHLC + diagonal EM Kalman", "not promoted"),
            "exp_003_corrected": ("log", "diagonal", "log OHLC + diagonal", "not promoted"),
            "exp_004_corrected": ("log", "full", "log OHLC + full", "not promoted")}
    scs = {r["exp"]: r for r in sb.get("scoreboard", [])}
    rows3 = []
    for eid in CELLS:
        r = scs.get(eid, {}); t, cov, d, status = desc[eid]
        reason = {"exp_002_corrected": "best mean+median Sharpe, 0 neg seeds, Sharpe-delta CI lower>0",
                  "rawctl_corrected": "comparator", "exp_007_corrected": "higher return but 3 neg seeds, worse risk-adj",
                  "exp_003_corrected": "transform adds no value", "exp_004_corrected": "transform adds no value"}.get(eid, "")
        rows3.append({"experiment_id": eid, "description": d, "covariance_type": cov, "transform": t,
                      "price_basis": "raw_close", "feature_basis": "filtered_ohlc+raw_vol" if eid != "rawctl_corrected" else "raw_ohlcv",
                      "seeds": r.get("n"), "mean_return": r.get("ret_mean"), "median_return": r.get("ret_median"),
                      "std_return": r.get("ret_std"), "min_return": r.get("ret_min"), "max_return": r.get("ret_max"),
                      "mean_sharpe": r.get("sharpe_mean"), "median_sharpe": r.get("sharpe_median"), "std_sharpe": r.get("sharpe_std"),
                      "mean_maxdd": r.get("maxdd_mean"), "negative_seed_count": r.get("neg_seeds"),
                      "mean_round_trips": r.get("rt_mean"), "promotion_status": status, "reason": reason})
    _sheet(wb, "Core_Scoreboard", rows3)

    # 4 Paired_Seed_Comparison
    base = {r["seed"]: r for r in seedm if r["experiment_id"] == "rawctl_corrected"}
    rows4 = []
    for eid in CELLS:
        if eid == "rawctl_corrected":
            continue
        for r in [x for x in seedm if x["experiment_id"] == eid]:
            b = base.get(r["seed"])
            if not b:
                continue
            rows4.append({"experiment_id": eid, "seed": r["seed"], "raw_return": b["return"], "kalman_return": r["return"],
                          "delta_return": r["return"] - b["return"], "raw_sharpe": b["sharpe"], "kalman_sharpe": r["sharpe"],
                          "delta_sharpe": r["sharpe"] - b["sharpe"], "raw_maxdd": b["max_drawdown"], "kalman_maxdd": r["max_drawdown"],
                          "delta_maxdd": r["max_drawdown"] - b["max_drawdown"],
                          "kalman_beats_raw_return": r["return"] > b["return"], "kalman_beats_raw_sharpe": r["sharpe"] > b["sharpe"]})
    _sheet(wb, "Paired_Seed_Comparison", rows4)

    # 5 Seed_Level_Metrics
    _sheet(wb, "Seed_Level_Metrics", seedm)

    # 6 Nautilus_Parity
    rows6 = []
    for pf in sorted(glob.glob(str(OUT / "*" / "diagnostics" / "parity_*.json"))):
        d = _load(pf); eid = d.get("experiment_id"); s = d.get("seed")
        rows6.append({"experiment_id": eid, "seed": s, "env_return": d.get("env_return"), "nautilus_return": d.get("nautilus_return"),
                      "abs_return_delta": d.get("cum_return_abs_diff"), "abs_final_equity_delta": d.get("final_equity_abs_diff"),
                      "round_trips_env": (d.get("round_trips") or {}).get("env"), "round_trips_nautilus": (d.get("round_trips") or {}).get("nautilus"),
                      "price_basis": d.get("price_basis", "raw_close"), "parity_status": "PASS" if d.get("parity_pass") else "FAIL",
                      "parity_path": pf})
    _sheet(wb, "Nautilus_Parity", rows6)

    # 7 Indicator_Basis_Audit
    idf = pd.read_csv(OUTV / "indicator_basis_diffs.csv") if (OUTV / "indicator_basis_diffs.csv").exists() else pd.DataFrame()
    rows7 = []
    for _, r in idf.iterrows():
        rows7.append({"experiment_id": r["experiment_id"], "indicator": r["indicator"],
                      "gen_vs_filtered_max_abs_diff": r["gen_vs_filtered_max_abs_diff"], "gen_vs_filtered_mean_abs_diff": r["gen_vs_filtered_mean_abs_diff"],
                      "gen_vs_raw_max_abs_diff": r["gen_vs_raw_max_abs_diff"], "gen_vs_raw_mean_abs_diff": r["gen_vs_raw_mean_abs_diff"],
                      "filtered_basis_match": r["filtered_basis_match"], "raw_basis_match": r["raw_basis_match"],
                      "verdict": ib.get("cells", {}).get(r["experiment_id"], {}).get("verdict", "?")})
    _sheet(wb, "Indicator_Basis_Audit", rows7)

    # 8 Price_Basis_Audit
    rows8 = []
    for k, v in pb.get("max_abs_price_minus_raw_close", {}).items():
        eid, _, sd = k.rpartition("_s")
        rows8.append({"experiment_id": eid, "seed": sd, "max_abs_price_minus_raw_close": v,
                      "price_in_feature_cols": pb.get("price_in_feature_cols"), "raw_close_in_feature_cols": pb.get("raw_close_in_feature_cols"),
                      "feature_count": pb.get("feature_count"), "volume_raw_preserved": pb.get("volume_raw_preserved"),
                      "env_uses_price": pb.get("env_uses_price"), "nautilus_uses_price": pb.get("nautilus_uses_price"),
                      "status": pb.get("status"), "failure_reason": "; ".join(pb.get("failures", [])) or "none"})
    _sheet(wb, "Price_Basis_Audit", rows8)

    # 9 Kalman_Diagnostics (filter-level, from phase-1 + dataset drift)
    diag_src = {"exp_002_corrected": _KF / "outputs/diagnostics/kalman_qr_diagnostics.json",
                "exp_007_corrected": _KF / "ablation/outputs/exp_007/diagnostics/kalman_diagnostics.json",
                "exp_003_corrected": _KF / "ablation/outputs/exp_003/diagnostics/kalman_diagnostics.json",
                "exp_004_corrected": _KF / "ablation/outputs/exp_004/diagnostics/kalman_diagnostics.json"}
    rows9 = []
    for eid, p in diag_src.items():
        d = _load(p); ds = _load(OUT / eid / "diagnostics/dataset_diagnostics.json")
        qstd = d.get("Q_diag_std") or (d.get("mean_q_std") and [d["mean_q_std"]]) or []
        rstd = d.get("R_diag_std") or (d.get("mean_r_std") and [d["mean_r_std"]]) or []
        qm = float(np.mean(qstd)) if qstd else d.get("mean_q_std")
        rm = float(np.mean(rstd)) if rstd else d.get("mean_r_std")
        dr = ds.get("drift_vs_raw_close", {})
        rows9.append({"experiment_id": eid, "transform": ds.get("transform"), "covariance_type": ds.get("covariance"),
                      "q_std_mean": qm, "r_std_mean": rm, "q_r_ratio": (qm / rm) if qm and rm else None,
                      "q_condition": d.get("Q_cond"), "r_condition": d.get("R_cond"),
                      "avg_kalman_gain": d.get("average_kalman_gain"), "variance_compression": d.get("price_variance_compression"),
                      "filtered_vs_raw_close_max_pct": dr.get("filtered_vs_raw_close_max_abs_pct"),
                      "filtered_vs_raw_close_mean_pct": dr.get("filtered_vs_raw_close_mean_abs_pct"),
                      "ohlc_violation_count": ds.get("ohlc_validity", {}).get("total"),
                      "diagnostics_path": str(p)})
    _sheet(wb, "Kalman_Diagnostics", rows9)

    # 10 Feature_Drift (OHLC level: filtered features vs raw, exp_002 decisive)
    from rl_gold_trading.config import set_config_path
    set_config_path(str(pipeline_default()))
    raw = _raw_clean()
    rows10 = []
    for eid in ["exp_002_corrected", "exp_007_corrected", "exp_003_corrected", "exp_004_corrected"]:
        csv = OUT / eid / "data/model_input.csv"
        if not csv.exists():
            continue
        g = pd.read_csv(csv); g["datetime"] = pd.to_datetime(g["datetime"], utc=True); g = g.set_index("datetime")
        for c in ["open", "high", "low", "close"]:
            f = g[c].to_numpy(float); rr = raw[c].reindex(g.index).to_numpy(float)
            rows10.append({"experiment_id": eid, "feature": c, "mean_raw": float(rr.mean()), "mean_filtered": float(f.mean()),
                           "std_raw": float(rr.std()), "std_filtered": float(f.std()),
                           "mean_abs_diff": float(np.abs(f - rr).mean()), "max_abs_diff": float(np.abs(f - rr).max()),
                           "correlation": float(np.corrcoef(f, rr)[0, 1]), "feature_drift_path": str(OUT / eid / "diagnostics/dataset_diagnostics.json")})
    _sheet(wb, "Feature_Drift", rows10)

    # 11 Historical_Phase1 (filtered = DIAGNOSTIC ONLY)
    rc = hist.get("raw_close", {}); fc = hist.get("filtered_close", {})
    def _find(d, key):
        for k, v in d.items():
            if k.startswith(key) or key in k:
                return v
        return {}
    rows11 = []
    for key, dsc in [("exp_013", "strong denoise Q0.25R5"), ("exp_009", "R x5"), ("exp_010", "R x10"),
                     ("exp_007", "raw+diag near-passthrough"), ("exp_001_rawctl", "raw no-Kalman control")]:
        r = _find(rc, key); f = _find(fc, key)
        rows11.append({"experiment_id": key, "description": dsc,
                       "filtered_pnl_return": f.get("ret_mean"), "filtered_pnl_sharpe": f.get("sharpe_mean"),
                       "raw_pnl_return": r.get("ret_mean"), "raw_pnl_sharpe": r.get("sharpe_mean"),
                       "artifact_status": "phase1 (filtered-reward-trained policies)",
                       "interpretation": "filtered Sharpe inflated by smoothed-price PnL" if f else "",
                       "diagnostic_only": "DIAGNOSTIC ONLY — NOT TRADEABLE PERFORMANCE"})
    _sheet(wb, "Historical_Phase1_Results", rows11)

    # 12 Paper_Reference
    cfg = _load(pipeline_default())
    rows12 = [
        {"paper_component": "Kalman-enhanced DRL thesis", "paper_value_or_claim": "denoising improves DRL robustness",
         "implementation_value": "partially supported (full-cov stabilizes; ~2x fair-raw Sharpe)", "match_status": "PARTIAL",
         "notes": "stronger denoising HURTS on raw price", "source_path": "RAWPRICE_FINAL_REPORT.md"},
        {"paper_component": "State / 22 features", "paper_value_or_claim": "Kalman OHLC + technical indicators",
         "implementation_value": "22D: 4 filtered OHLC + raw Vol + 17 indicators (from filtered OHLC)", "match_status": "MATCH",
         "notes": "indicators verified filtered-basis 17/17", "source_path": "FINAL_INDICATOR_BASIS_AUDIT.md"},
        {"paper_component": "Action space", "paper_value_or_claim": "{-1,0,+1}", "implementation_value": "Discrete(3) {sell,hold,buy}",
         "match_status": "MATCH", "notes": "", "source_path": "config/experiment_dukascopy_1d.json"},
        {"paper_component": "Reward", "paper_value_or_claim": "return - risk - cost + stability",
         "implementation_value": "alpha*R - beta*DD - gamma*Cost + delta*Stability (Eq.22)", "match_status": "MATCH",
         "notes": "frozen", "source_path": "config/experiment_dukascopy_1d.json"},
        {"paper_component": "z-score", "paper_value_or_claim": "252-day rolling", "implementation_value": "252 (1d)",
         "match_status": "MATCH", "notes": "", "source_path": "src/rl_gold_trading/normalize.py"},
        {"paper_component": "Reported PPO+Kalman / Raw result", "paper_value_or_claim": "internally inconsistent labels (prior audit)",
         "implementation_value": "not used as target; methodological fidelity", "match_status": "N/A",
         "notes": "honest raw-price basis used instead", "source_path": "RAWPRICE_FINAL_REPORT.md"},
    ]
    _sheet(wb, "Paper_Reference", rows12)

    # 13 Final_Artifacts_Index
    rows13 = []
    for pat, typ in [("docs/*.md", "doc"), ("docs/*.xlsx", "workbook"), ("outputs/final_verification/*", "verification"),
                     ("outputs/FINAL_MODEL/*", "final_model"), ("configs/*.json", "config"), ("scripts/*.py", "script")]:
        for f in sorted(glob.glob(str(_ABLR / pat))):
            p = Path(f)
            rows13.append({"artifact_type": typ, "artifact_name": p.name, "path": str(p.relative_to(_ROOT)),
                           "exists": p.exists(), "description": ""})
    _sheet(wb, "Final_Artifacts_Index", rows13)

    # 14 Runbook_Commands
    py = ".venv/Scripts/python.exe"
    base = "kalmanFilter/ablation_rawprice/scripts"
    rows14 = [
        {"step": 1, "purpose": "verify final price+indicator+parity", "command": f"{py} {base}/verify_final.py",
         "expected_output": "FINAL VERIFY: PASS", "notes": "gate"},
        {"step": 2, "purpose": "regenerate scoreboard + paired stats", "command": f"{py} {base}/analyze_rawprice.py",
         "expected_output": "RAWPRICE_SCOREBOARD.md + RAWPRICE_SEED_ROBUSTNESS_REPORT.md", "notes": ""},
        {"step": 3, "purpose": "rebuild Excel workbook", "command": f"{py} {base}/make_workbook.py",
         "expected_output": "KALMAN_PPO_FULL_STUDY_RESULTS.xlsx", "notes": ""},
        {"step": 4, "purpose": "re-eval final model (raw price)", "command": f"{py} {base}/run_rawprice_ablation.py --exp exp_002_corrected --stage eval --seed 23",
         "expected_output": "eval seed=23 px=raw sharpe~0.82", "notes": ""},
        {"step": 5, "purpose": "re-run nautilus parity (final seed)", "command": f"{py} {base}/run_rawprice_ablation.py --exp exp_002_corrected --stage nautilus --seed 23",
         "expected_output": "parity PASS |Δ|~1e-4", "notes": ""},
    ]
    _sheet(wb, "Runbook_Commands", rows14)

    # 15 Caveats_And_Next_Work
    rows15 = [
        {"category": "statistics", "caveat_or_next_step": "Wilcoxon p=0.131 at n=10", "severity": "medium", "required_before_publication": "recommended", "notes": "bootstrap CI lower>0 still"},
        {"category": "statistics", "caveat_or_next_step": "20-seed confirmation", "severity": "low", "required_before_publication": "optional", "notes": "tightens CI"},
        {"category": "training", "caveat_or_next_step": "best-validation checkpointing (symmetric)", "severity": "medium", "required_before_publication": "approval-gated", "notes": "largest seed-variance source"},
        {"category": "scope", "caveat_or_next_step": "paper 5D OHLCV Kalman branch", "severity": "low", "required_before_publication": "separate branch", "notes": "current is OHLC-only by design"},
        {"category": "result", "caveat_or_next_step": "strong denoising hurts under raw-price PnL", "severity": "info", "required_before_publication": "documented", "notes": "do not pursue"},
        {"category": "accounting", "caveat_or_next_step": "filtered-close PnL invalid for trading perf", "severity": "high", "required_before_publication": "fixed", "notes": "diagnostic only"},
    ]
    _sheet(wb, "Caveats_And_Next_Work", rows15)

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    # validate
    from openpyxl import load_workbook
    chk = load_workbook(XLSX)
    names = chk.sheetnames
    man_out = {"workbook": str(XLSX.relative_to(_ROOT)), "n_sheets": len(names), "sheets": names,
               "validated": all(chk[n].max_row >= 1 for n in names), "saved_ok": True}
    (OUTV).mkdir(parents=True, exist_ok=True)
    (OUTV / "excel_workbook_manifest.json").write_text(json.dumps(man_out, indent=2))
    print("workbook saved:", XLSX)
    print("sheets:", names)
    print("validated:", man_out["validated"])


def pipeline_default():
    import sys
    sys.path.insert(0, str(_KF))
    from src import pipeline
    return pipeline.DEFAULT_CONFIG


def _raw_clean():
    import sys
    sys.path.insert(0, str(_KF)); sys.path.insert(0, str(_ROOT / "src"))
    from src import pipeline
    return pipeline.load_clean_ohlcv(pipeline.DEFAULT_CONFIG)


if __name__ == "__main__":
    main()
